# import libraries
import urllib.request
from bs4 import BeautifulSoup
import openpyxl as op
import os
import requests
from hashlib import md5
import random

# change trait and path
trait="cirrhosis"
path="D:/document/bioInfo/gwasID/"
isGwasInfo=False

# params
url = 'https://gwas.mrcieu.ac.uk/datasets/?gwas_id__icontains=&year__iexact=&trait__icontains='+trait+'&consortium__icontains='
rows = []

# baidu api translate
appid = '20231130001896707'
appkey = 'MkW4prMH1Nk2XI8dYdOY'
from_lang = 'en'
to_lang = 'zh'
endpoint = 'http://api.fanyi.baidu.com'
baiduPath = '/api/trans/vip/translate'
baiduUrl = endpoint + baiduPath

def make_md5(s, encoding='utf-8'):
    return md5(s.encode(encoding)).hexdigest()

def baidu_api(query, from_lang, to_lang):
    salt = random.randint(32768, 65536)
    sign = make_md5(appid + query + str(salt) + appkey)
    headers = {'Content-Type': 'application/x-www-form-urlencoded'}
    payload = {'appid': appid, 'q': query, 'from': from_lang, 'to': to_lang, 'salt': salt, 'sign': sign}
    r = requests.post(baiduUrl, params=payload, headers=headers)
    result = r.json()
    return result["trans_result"][0]['dst']

# get GWAS id
def getRows(url):
    while (True):
        try:
            page = urllib.request.urlopen(url, timeout=10).read().decode('utf-8')
        except Exception as error:
            print('Error:', error)
        else:
            break
    soup = BeautifulSoup(page, 'html.parser')
    table = soup.find('table')
    results = table.find_all('td', class_="text-nowrap")
    if (len(results) == 0):
        print("no available data")
    else:
        if(isGwasInfo):

            if(os.path.exists(path+trait+'.xlsx')):
                wb = op.load_workbook(filename=path+trait+'.xlsx')
                sheet = wb.worksheets[0]
                row = sheet.max_row+1
            else:
                wb = op.Workbook()
                sheet = wb.active
                sheet.cell(row=1, column=1, value="id")
                sheet.cell(row=1, column=2, value="population")
                sheet.cell(row=1, column=3, value="sample_size")
                sheet.cell(row=1, column=4, value="number_of_snps")
                sheet.cell(row=1, column=5, value="year")
                sheet.cell(row=1, column=6, value="pmid")
                sheet.cell(row=1, column=7, value="consortium")
                sheet.cell(row=1, column=8, value="trait")
                row = 2

            for result in results:
                data = result.string
                print(data)
                info = getInfomation(data)
                sheet.cell(row=row, column=1, value=data)
                for j in range(0, len(info), 1):
                    sheet.cell(row=row, column=j+2, value=info[j])
                row += 1
            wb.save(path+trait+'.xlsx')
        else:
            for result in results:
                data = result.string
                print(data)
                rows.append(data)
            # export
            with open(path + trait + ".txt", "w") as file:
                for line in rows:
                    file.write(line + "\n")


def getInfomation(id):
    url = 'https://gwas.mrcieu.ac.uk/datasets/' + id + '/'
    while(True):
        try:
            page = urllib.request.urlopen(url, timeout=10).read().decode('utf-8')
        except Exception as error:
            print('Error:',error,' ',url)
        else:
            break
    soup = BeautifulSoup(page, 'html.parser')
    trait=soup.find_all("h1")[0].string
    trait= trait+"||"+baidu_api(trait,from_lang,to_lang)
    pages = soup.find_all('th', class_='text-nowrap')
    population='NA'
    pmid='NA'
    year='NA'
    sample_size='NA'
    nsp='NA'
    consortium='NA'
    for i in range(0, len(pages), 1):
        temp = pages[i]

        if (temp.text == "PMID"):
            pmid = temp.findNext("td").text
            continue
        if (temp.text == "Year"):
            year = temp.findNext("td").text
            continue
        if (temp.text == "Population"):
            population = temp.findNext("td").text
            continue
        if (temp.text == "Sample size"):
            sample_size = temp.findNext("td").text
            continue
        if (temp.text == "Number of SNPs"):
            nsp = temp.findNext("td").text
            continue
        if (temp.text == "Consortium"):
            consortium = temp.findNext("td").text
            continue
    result = [population, sample_size, nsp, year, pmid, consortium,trait]
    return result

# fill rows
while (True):
    try:
        page = urllib.request.urlopen(url, timeout=10).read().decode('utf-8')
    except Exception as error:
        print('Error:', error)
    else:
        break
soup = BeautifulSoup(page, 'html.parser')
pages = soup.find_all('a', class_='page-link')
if (len(pages) <= 2):
    getRows(url)
else:
    maxPage = pages[len(pages) - 2].contents[0].replace("\n", "").replace(" ", "")
    maxPage = int(maxPage)
    for pageCout in range(1, maxPage + 1, 1):
        urlpage = url + '&page=' + str(pageCout)
        getRows(urlpage)
print(trait + " created")