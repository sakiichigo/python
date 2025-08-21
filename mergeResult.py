import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# 指定根目录
root_dir = r"D:\document\bioInfo\covid-alps\result"

# 创建一个新的工作簿
output_workbook = Workbook()
# 删除默认创建的工作表
if "Sheet" in output_workbook.sheetnames:
    output_workbook.remove(output_workbook["Sheet"])

# 遍历根目录下的所有子文件夹
for folder_name in os.listdir(root_dir):
    folder_path = os.path.join(root_dir, folder_name)

    # 检查是否为文件夹
    if os.path.isdir(folder_path):
        # 创建以文件夹命名的工作表
        sheet = output_workbook.create_sheet(title=folder_name)
        header_saved = False  # 标记是否已保存表头

        # 遍历文件夹中的所有文件
        for file_name in os.listdir(folder_path):
            file_path = os.path.join(folder_path, file_name)

            # 检查文件类型
            if file_name.endswith('.xlsx'):
                file_type = 'xlsx'
            elif file_name.endswith('.csv'):
                file_type = 'csv'
            else:
                print(f"跳过非Excel/CSV文件: {file_path}")
                continue

            try:
                # 读取文件（保留表头）
                if file_type == 'xlsx':
                    df = pd.read_excel(file_path, header=0)  # 第一行为表头
                else:  # CSV文件
                    try:
                        df = pd.read_csv(file_path, sep=',', header=0)
                        if len(df.columns) == 1:  # 尝试制表符分隔
                            df = pd.read_csv(file_path, sep='\t', header=0)
                    except UnicodeDecodeError:
                        df = pd.read_csv(file_path, sep=',', encoding='latin-1', header=0)
                        if len(df.columns) == 1:
                            df = pd.read_csv(file_path, sep='\t', encoding='latin-1', header=0)

                # 处理表头和数据
                if not header_saved:
                    # 第一个文件：保留表头和数据
                    for cell in dataframe_to_rows(df, index=False, header=True):
                        sheet.append(cell)
                    header_saved = True  # 标记表头已保存
                else:
                    # 后续文件：只保留数据（跳过表头）
                    for cell in dataframe_to_rows(df, index=False, header=False):
                        sheet.append(cell)

                print(f"已合并: {file_path} 到工作表: {folder_name}")
            except Exception as e:
                print(f"无法合并 {file_path}: {str(e)}")

# 保存合并后的Excel文件
output_file = os.path.join(root_dir, "合并结果.xlsx")
output_workbook.save(output_file)
print(f"合并完成，文件已保存至: {output_file}")