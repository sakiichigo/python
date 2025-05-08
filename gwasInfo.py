# import libraries
import urllib.request
from bs4 import BeautifulSoup
import openpyxl as op
import socket
import requests
import random
import json
from hashlib import md5

#筛选过程，对获取的excel结果表填充人口、样本量、nsps、年份信息

# change excelUrl
excelUrl = "C:/Users/user/Desktop/test1.xlsx"

# get population,sample_size,nsp,year
def getInfomation(id):
    url = 'https://gwas.mrcieu.ac.uk/datasets/' + id + '/'
    while(True):
        try:
            page = urllib.request.urlopen(url, timeout=10).read().decode('utf-8')
        except Exception as error:
            print('Error:',error,' ',url)
        else:
            break
    soup = BeautifulSoup(page, 'html.parser')
    pages = soup.find_all('th', class_='text-nowrap')
    population = "Nome"
    sample_size = "Nome"
    nsp = "Nome"
    year = "Nome"
    pmid = "Nome"
    consortium = "Nome"
    for i in range(0,len(pages),1):
        temp=pages[i]
        if (temp.text == "PMID"):
            pmid = temp.findNext("td").text
            continue
        if(temp.text=="Year"):
            year = temp.findNext("td").text
            continue
        if(temp.text=="Population"):
            population = temp.findNext("td").text
            continue
        if (temp.text == "Sample size"):
            sample_size = temp.findNext("td").text
            continue
        if (temp.text == "Number of SNPs"):
            nsp = temp.findNext("td").text
            continue
        if (temp.text == "Consortium"):
            consortium = temp.findNext("td").text
            continue
    result=[population,sample_size,nsp,year,pmid,consortium]
    return result

# baidu api translate
appid = '20231130001896707'
appkey = 'MkW4prMH1Nk2XI8dYdOY'
from_lang = 'en'
to_lang = 'zh'
endpoint = 'http://api.fanyi.baidu.com'
path = '/api/trans/vip/translate'
url = endpoint + path

# Generate salt and sign
def make_md5(s, encoding='utf-8'):
    return md5(s.encode(encoding)).hexdigest()

def baidu_api(query, from_lang, to_lang):
    salt = random.randint(32768, 65536)
    sign = make_md5(appid + query + str(salt) + appkey)
    headers = {'Content-Type': 'application/x-www-form-urlencoded'}
    payload = {'appid': appid, 'q': query, 'from': from_lang, 'to': to_lang, 'salt': salt, 'sign': sign}
    r = requests.post(url, params=payload, headers=headers)
    result = r.json()
    return result["trans_result"][0]['dst']

# check element
def checkElement(check,row,column):
    if(sheet.cell(row=row, column=column).value==None):
        sheet.insert_cols(column)
    sheet.cell(row=row, column=column, value=check)

# modify xlsx
wb = op.load_workbook(excelUrl)
for i in range(0,len(wb.worksheets),1):
    sheet = wb.worksheets[i]
    # 获取最后一列的列号并加1，作为新的起始列
    startIndex = sheet.max_column + 1
    checkElement('population', 1, startIndex)
    checkElement('sample_size', 1, startIndex+1)
    checkElement('number_of_snps', 1, startIndex+2)
    checkElement('year', 1, startIndex+3)
    checkElement('exposure_zh', 1, startIndex+4)
    checkElement('outcome_zh', 1, startIndex+5)
    checkElement('pmid', 1, startIndex+6)
    checkElement('Consortium', 1, startIndex+7)
    for j in range(2,sheet.max_row+1,1):

        exposure=sheet.cell(row=j, column=1).value
        outcome=sheet.cell(row=j, column=2).value
        outcome_en = sheet.cell(row=j, column=3).value
        exposure_en = sheet.cell(row=j, column=4).value
        if(exposure==None and outcome==None):
            continue

        if(exposure!=sheet.cell(row=j-1, column=1).value):
            if (exposure_en != None):
                exposure_zh=baidu_api(exposure_en.split("||")[0].replace("_"," ").replace("-"," "),from_lang, to_lang)
            else:
                exposure_zh = None
            if (exposure != None):
                inf_exposure = getInfomation(exposure)
            else:
                inf_exposure = None
        if(outcome!=sheet.cell(row=j-1, column=2).value):
            if (outcome_en != None):
                outcome_zh = baidu_api(outcome_en.split("||")[0].replace("_"," ").replace("-"," "), from_lang, to_lang)
            else:
                outcome_zh = None
            if (outcome != None):
                inf_outcome = getInfomation(outcome)
            else:
                inf_outcome = None

        #population,sample_size,nsp,year,pmid,consortium
        if(outcome!=None and exposure!= None):
            print("exposure|outcome population :" + inf_exposure[0] + "|" + inf_outcome[0])
            if (inf_exposure[0] == inf_outcome[0]):
                sheet.cell(row=j, column=startIndex, value=inf_exposure[0])
            sheet.cell(row=j, column=startIndex+1, value=str(inf_exposure[1]) + "|" + inf_outcome[1])
            sheet.cell(row=j, column=startIndex+2, value=str(inf_exposure[2]) + "|" + inf_outcome[2])
            sheet.cell(row=j, column=startIndex+3, value=str(inf_exposure[3]) + "|" + inf_outcome[3])
            sheet.cell(row=j, column=startIndex+4, value=exposure_zh)
            sheet.cell(row=j, column=startIndex+5, value=outcome_zh)
            sheet.cell(row=j, column=startIndex+6, value=str(inf_exposure[4]) + "|" + inf_outcome[4])
            sheet.cell(row=j, column=startIndex+7, value=str(inf_exposure[5]) + "|" + inf_outcome[5])
        elif(exposure!= None):
            print("exposure population :" + inf_exposure[0])
            sheet.cell(row=j, column=startIndex, value=inf_exposure[0])
            sheet.cell(row=j, column=startIndex+1, value=str(inf_exposure[1]))
            sheet.cell(row=j, column=startIndex+2, value=str(inf_exposure[2]))
            sheet.cell(row=j, column=startIndex+3, value=str(inf_exposure[3]))
            sheet.cell(row=j, column=startIndex+4, value=exposure_zh)
            sheet.cell(row=j, column=startIndex+5, value=str(inf_exposure[4]))
            sheet.cell(row=j, column=startIndex+6, value=str(inf_exposure[5]))

    wb.save(excelUrl)
    print(wb.worksheets[i].title+" saved")